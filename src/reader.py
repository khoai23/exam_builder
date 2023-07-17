"""
Basic code to read data from a csv file.
"""
import os, io, csv 
import glob
import traceback
import shutil
import string
import openpyxl, openpyxl_image_loader 
import base64

from src.image import DefaultClient, check_and_write_image
from src.utils import cell_name_xl_to_tuple

import logging
logger = logging.getLogger(__name__)

from typing import Optional, List, Tuple, Any, Union, Dict

_DEFAULT_FILE_PREFIX = "test/sample"
_DEFAULT_BACKUP_PREFIX = "test/backup"

TEMPORARY_FILE_DIR = "test"
_DEFAULT_RECOVER_FILE_PREFIX = "test/recover_current"
_DEFAULT_RECOVER_BACKUP_PREFIX = "test/recover_backup"
# no longer hardfix - now try to load file by a prefix, xlsx first
# DEFAULT_FILE_PATH = "test/sample.xlsx"
def get_file_by_prefix(prefix: str, prefer_cue: Optional[str]=None):
    valid_file = glob.glob(prefix + "*")
    if(len(valid_file) == 0):
        raise FileNotFoundError("Cannot find file of prefix {}. Check your input".format(prefix))
    if(prefer_cue):
        for f in valid_file:
            if(prefer_cue in f):
                return f 
    return valid_file[0]
DEFAULT_FILE_PATH = get_file_by_prefix(_DEFAULT_FILE_PREFIX, prefer_cue=".xlsx")
try:
    DEFAULT_BACKUP_PATH = get_file_by_prefix(_DEFAULT_BACKUP_PREFIX, prefer_cue=".xlsx")
except FileNotFoundError:
    DEFAULT_BACKUP_PATH = None

def move_file(source: str, target: str, is_target_prefix: Optional[bool]=True, autoremove_target: bool=True):
    # move file. is_target_prefix=true will have target appending the extension of the source 
    # return the location of the target 
    if(is_target_prefix):
        _, ext = os.path.splitext(source)
        target = target + ext 
    if(os.path.isfile(target) and autoremove_target):
        os.remove(target)
    shutil.move(source, target)
    return target

def copy_file(source: str, target: str, is_target_prefix: Optional[bool]=True, autoremove_target: bool=True):
    # copy file. is_target_prefix=true will have target appending the extension of the source 
    # return the location of the target 
    # TODO merge with move?
    if(is_target_prefix):
        _, ext = os.path.splitext(source)
        target = target + ext 
    if(os.path.isfile(target) and autoremove_target):
        os.remove(target)
    shutil.copy(source, target)
    return target

HEADERS = ["question", "answer1", "answer2", "answer3", "answer4", "correct_id", "category", "tag", "special", "variable_limitation"]
SPECIAL_TAGS = ["is_multiple_choice", "is_dynamic_key", "is_fixed_equation", "is_single_equation", "is_single_option"]

def read_file(filepath: str, headers: Optional[List[str]]=None, strict: bool=False):
    if(".csv" in filepath):
        return read_file_csv(filepath, headers=headers, strict=strict)
    elif(".xlsx" in filepath):
        return read_file_xlsx(filepath, headers=headers, strict=strict)
    else:
        raise ValueError("Unusable filepath: {}; check the extension (must be .csv/.xlsx)".format(filepath))

def read_file_csv(filepath: str, headers: Optional[List[str]]=None, strict: bool=False, ignore_failed_row: bool=False):
    # read a file from `filepath` into a dict. Should at minimum has `question`, `answer1-4`, and `correct_id`
    data = []
    if(ignore_failed_row):
        failed_rows = []
    with io.open(filepath, "r", encoding="utf-8") as rf:
        reader = csv.DictReader(rf, fieldnames=headers)
        for i, row in enumerate(reader):
            try:
                data.append(process_field(row))
            except Exception as e:
                if(ignore_failed_row):
                    # logger.error("Failed row: {}".format(e, traceback.format_exc()))
                    failed_rows.append(i+1)
                    continue
                else:
                    raise e
    if(strict):
        fields = ("question", "correct_id", "answer1", "answer2", "answer3", "answer4")
        valid_data = lambda row: all(field in row for field in fields) or row["is_single_equation"]
        assert all(valid_data(row) for row in data), "Read data missing field; exiting: {}".format(data)
    if(ignore_failed_row):
        return failed_rows, data
    else:
        return data

def read_file_xlsx(filepath: str, headers: Optional[List[str]]=None, strict: bool=False, ignore_failed_row: bool=False):
    workbook = openpyxl.load_workbook(filepath)
    data_sheet = workbook[workbook.sheetnames[0]]
    logger.debug("Expecting the first sheet to contain the data; which is: {}".format(workbook.sheetnames[0]))
    data = []
    if(ignore_failed_row):
        failed_rows = []
    image_dictionary = dict()
    if(DefaultClient is None):
        logger.warning("Image module disabled/uninitialized. Import with images is disabled.")
    else:
        image_loader = openpyxl_image_loader.SheetImageLoader(data_sheet)
        # old images will be read written in as |||{image_key}|||; b64 format
        # new images will be read and converted to same format as above
    #    logger.debug(image_loader._images)
    #    image_loader.get("B2").show()
        for key in image_loader._images:
            logger.debug("Converting image for cell {}, exist: {}".format(key, image_loader.image_in(key)))
            try:
                img_buffer = io.BytesIO()
                image_loader.get(key).save(img_buffer, format="PNG")
            except ValueError:
                # due to some weird bug, subsequent file opening can have old references of other loader. It will output "I/O operation on closed file" if not checked 
                # hence, when receiving this, simply ignore them 
                logger.info("Image at cell {} cannot be read. Ignoring.".format(key))
                continue
            img_data = img_buffer.getvalue()
            number_key = row, col = cell_name_xl_to_tuple(key)
            # logger.debug("Image at {:s} - {}: {}...".format(key, cell_name_xl_to_tuple(key), data[:100]))
            row_dict = image_dictionary[row] = image_dictionary.get(row, dict())
            # col is actually corresponding to answer1-4 already, lucky!
            row_dict[col] = check_and_write_image(img_data)
    for i, row in enumerate(data_sheet.iter_rows(values_only=True)):
        if(i == 0 and headers is None):
            # if header is None, load it from the excel 
            headers = row;
            continue 
        # only add the data field in when value is not empty 
        # have to re-convert back to string right now; TODO code to skip this
        try:
            data.append(process_field({header: str(value) for header, value in zip(headers, row) if value is not None }, image_dictionary=image_dictionary.get(i, None)))
        except Exception as e:
            if(ignore_failed_row):
                # logger.error("Failed row: {}".format(e, traceback.format_exc()))
                failed_rows.append(i+1)
                continue
            else:
                raise e
    if(strict):
        fields = ("question", "correct_id", "answer1", "answer2", "answer3", "answer4")
        valid_data = lambda row: all(field in row for field in fields)
        assert all(valid_data(row) for row in data), "Read data missing field; exiting: {}".format(data)
    if(ignore_failed_row):
        return failed_rows, data
    else:
        return data

def write_file_csv(filepath: str, data: List[Dict], headers: List[str]=HEADERS):
    # write a csv file using the current read data. Useful when importing in append mode
    # newline to prevent duplicate writing
    with io.open(filepath, "w", encoding="utf-8", newline="") as wf:
        writer = csv.DictWriter(wf, fieldnames=headers)
        writer.writeheader()
        for row in data:
            writer.writerow(reconvert_field(row))
    return filepath

def write_file_xlsx(filepath: str, data: List[Dict], headers: List[str]=HEADERS):
    # write a xlsx file using the current read data.
    # TODO write the help sheet to it as well 
    # TODO keep formatting
    workbook = openpyxl.Workbook()
    data_sheet = workbook.create_sheet("QuestionSheet", 0)
    # write header 
    data_sheet.append(headers)
    # write data 
    for row in data:
        row = reconvert_field(row)
        data_row = [row.get(h, None) for h in headers ]
        data_sheet.append(data_row)
    # save 
    workbook.save(filepath)
    return filepath


def reconvert_field(row: Dict):
    """Reconvert data row back into valid writable in csv/xlsx here.
    MUST CREATE NEW DICT from this to prevent backporting into good data."""
    row = dict(row)
    special = [s for s in SPECIAL_TAGS if row.pop(s, None)]
    row["special"] = ", ".join(special)
    # reconvert tag & correct_id to corresponding format if necessary
    if("tag" in row):
        row["tag"] = ", ".join(row["tag"])
    if(isinstance(row["correct_id"], tuple)):
        row["correct_id"] = ", ".join([str(i) for i in row["correct_id"]])
    return row

def process_field(row, lowercase_field: bool=True, delimiter: str=",", image_dictionary: Optional[Dict]=None, image_format: str="|||{}|||"):
    """All processing of fields is done here.
    image_dictionary: if supplied, is reading & converting from external xlsx; image is already uploaded and can be referred by link
    image_format: image links will be wrapped by this update
    """
    new_data = {"is_multiple_choice": False}
    for k, v in row.items():
        if(v is None):
            continue
        v = v.strip()
        if(lowercase_field):
            k = k.lower()
        if(k == "tag"):
            # for tag field, split it by delimiter 
            v = [] if v == "" else [v.strip()] if delimiter not in v else [t.strip() for t in v.split(delimiter)]
#            logger.debug("Tag: ", v)
            # backward compatibility
            for st in SPECIAL_TAGS:
                if(st in v):
                    v.remove(st)
                    new_data[st] = True
        if(k == "special"):
            # if has tag for multiple choice, swap it to is_multiple_choice
            # TODO enforce mutual exclusivity
            for st in SPECIAL_TAGS:
                if(st in v):
                    new_data[st] = True
        if(k == "correct_id"): 
            try:
                if("," in v):
                    new_data["is_multiple_choice"] = True 
                    v = tuple(int(iv) for iv in v.split(","))
                    assert all(( 0 < iv <= 4 for iv in v)), "Correct_id is fixed to [1, 4] for now, but received: {}".format(v)
                else:
                    v = int(v)
                    assert 0 < v <= 4, "Correct_id is fixed to [1, 4] for now, but received: {}".format(v)
            except ValueError as e:
                logger.debug("The correct id `{}` cannot be parsed".format(v))
                raise e
        new_data[k] = v
    # If there is an image present in cell, replace it wholesale. TODO allow insertion 
    if(image_dictionary):
        for col, image_info in image_dictionary.items():
            if(0 < col <= 4):
                # image IS cell; replace directly
                new_data["answer{}".format(col)] = image_format.format(image_info["link"])
            elif(10 <= col < 16):
                # image is referred by cell, find and replace in question/answer1-4 
                # image is going from 1 - 6
                cue = "{image_" + (col - 10 + 1) + "}"
                used = False
                for field in ("question", "answer1", "answer2", "answer3", "answer4"):
                    if cue in new_data.get(field, ""): # preventing issue with is_single_equation variant
                        new_data[field] = new_data[field].replace(cue, image_format.format(image_info["link"]))
                        used = True 
                if not used:
                    # print warning that image is not used anywhere.
                    logger.warning("Image " + image_info["link"] + " has not been used; make sure to have a correct reference as {" + "image_" + (col - 10 + 1) + "} in either question or answer")
            else:
                # image is out of expected column, ignore
                logger.warning("Image {} is at invalid column {} (should be put at image_1-image6/10-15); not used.".format(image_info["link"], col))
    # assert no duplicate answers.
    answers = [v for k, v in new_data.items() if "answer" in k]
    assert "question" in new_data, "Data must have a valid question field."
    # fixed equation only has answer1; TODO if there is answer2/3/4 then fire a warning
    assert new_data.get("is_single_equation", False) or new_data.get("is_single_option", False) or \
        len(set(answers)) == len(answers), "There are duplicates in the list of answers of: {}".format(new_data)
    # if is_single_option, make sure that it has at least 4 corresponding templates.
    if(new_data.get("is_single_option", False) and new_data["variable_limitation"].count("\n") < 4):
        raise ValueError("Question {} must have at least 4 variant, but only has {}".format(new_data, new_data["variable_limitation"].count("\n")+1))
    # if multiple-choice question with only a single selection, convert it to list 
    if(new_data["is_multiple_choice"] and isinstance(new_data["correct_id"], int)):
        new_data["correct_id"] = (new_data["correct_id"],)
    return new_data

if __name__ == "__main__":
    # normal test
    data = read_file("test/sample.xlsx")
    logger.debug(data)
##    write_file_csv("test/test_write_sample.csv", data, HEADERS)
#    write_file_xlsx("test/test_write_sample.xlsx", data, HEADERS)
#    logger.debug("Tested writing.")
    # image read test
#    logger.debug(read_file_xlsx("test/sample_with_image.xlsx"))
